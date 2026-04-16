
import logging
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import LoginForm
    # Get principal's school
from schools.models import Principal, ReportingYear
from waste_management.models import WasteEntry
from django.db.models import Sum
from decimal import Decimal
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import login, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .utils import is_token_valid
from schools.models import School, ReportingYear
from recycler_management.models import Recycler
from accounts.models import User
from waste_management.models import WasteEntry
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from decimal import Decimal
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.db.models.functions import Coalesce



logger = logging.getLogger(__name__)

def login_view(request):
    if request.user.is_authenticated:
        logger.debug(f"[login_view] Already logged in: {request.user.username}")
        return redirect_by_role(request.user)

    form = LoginForm()

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        logger.debug(f"[login_view] Login attempt → username: {username}")

        # STEP 1 — ESG DB check
        user = authenticate(request, username=username, password=password)
        
        print(f"DEBUG STEP1: authenticate result = {user}")  # ← ADD

        if user is not None:
            if user.is_active:
                login(request, user)
                if user.must_change_password:
                    messages.warning(request, 'Please change your password before continuing.')
                    return redirect('change_password')
                return redirect_by_role(user)
            else:
                messages.error(request, 'Your account is inactive. Contact admin.')

        else:
            # STEP 2 — Evolvu API
            print(f"DEBUG STEP2: Going to Evolvu API for → {username}")  # ← ADD
            
            try:
                import requests as http_requests
                
                print(f"DEBUG STEP3: Calling Evolvu API...")  # ← ADD
                
                response = http_requests.post(
                    "https://sms.evolvu.in/arnolds_test/public/api/login",
                    json={"user_id": username, "password": password},
                    timeout=10
                )
                
                print(f"DEBUG STEP4: API status code = {response.status_code}")  # ← ADD
                print(f"DEBUG STEP5: API response = {response.text[:200]}")      # ← ADD
                
                data = response.json()
                
                print(f"DEBUG STEP6: token present = {bool(data.get('token'))}")  # ← ADD

            except Exception as e:
                print(f"DEBUG ERROR: Evolvu API failed → {e}")  # ← ADD
                messages.error(request, 'Invalid username or password.')
                return render(request, 'accounts/login.html', {'form': form})

            if data.get('token'):
                user_info = data.get('user', {})
                role_id   = user_info.get('role_id', '')
                user_details = data.get('userdetails', {})
                api_school_name = user_details.get('school_name', '')
                api_short_name = user_details.get('short_name', '')
                api_academic_year = user_details.get('academic_year', '')
                api_school_id = user_details.get('settings', {}).get('school_id', '')
                
                print(f"DEBUG STEP7: role_id = {role_id}, name = {user_info.get('name')}")  # ← ADD

                if role_id != 'T':
                    messages.error(request, 'Invalid username or password.')
                    return render(request, 'accounts/login.html', {'form': form})

                from accounts.models import User
                teacher = User.objects.filter(username=username).first()
                
                print(f"DEBUG STEP8: teacher in ESG DB = {teacher}")  # ← ADD

                if not teacher:
                    teacher = User.objects.create_user(
                        username=username,
                        password=password,
                        name=user_info.get('name', username),
                        role='teacher',
                        is_active=True,
                        school_id=api_school_id,
                        school_name=api_school_name,
                        school_short_name=api_short_name,
                        academic_year=api_academic_year,
                    )
                    print(f"DEBUG STEP9: New teacher created → {username}")  # ← ADD
                else:
                    teacher.name = user_info.get('name', teacher.name)
                    teacher.set_password(password)
                    teacher.school_id = api_school_id
                    teacher.school_name = api_school_name
                    teacher.school_short_name = api_short_name
                    teacher.academic_year = api_academic_year
                    teacher.save()
                    print(f"DEBUG STEP9: Teacher synced → {username}")  # ← ADD

                login(request, teacher)
                print(f"DEBUG STEP10: Login SUCCESS → redirecting to teacher_dashboard")  # ← ADD
                return redirect('teacher_dashboard')

            else:
                print(f"DEBUG: Token nahi aaya → {data}")  # ← ADD
                messages.error(request, 'Invalid username or password.')

    return render(request, 'accounts/login.html', {'form': form})

# =============================================================================
# LOGOUT VIEW
# =============================================================================
def logout_view(request):
    username = request.user.username
    logout(request)
    logger.debug(f"[logout_view] Logout → username: {username}")
    return redirect('login')


# =============================================================================
# ROLE BASED REDIRECT HELPER
# Role dekh ke sahi dashboard pe bhejo
# =============================================================================
def redirect_by_role(user):

    logger.debug(f"[redirect_by_role] Role: {user.role} → redirecting...")

    if user.role == 'superadmin':
        return redirect('superadmin_dashboard')      # ← fix: underscore added

    elif user.role == 'principal':
        return redirect('principal_dashboard')

    elif user.role == 'teacher':
        return redirect('teacher_dashboard')

    elif user.role == 'project_head':
        return redirect('projecthead_dashboard')

    elif user.role == 'recycler_admin':
        return redirect('recycler_dashboard')

    else:
        # Unknown role — wapas login pe
        logger.error(f"[redirect_by_role] Unknown role: {user.role} | Sending back to login")
        return redirect('login')



# @login_required(login_url='login')
# def superadmin_dashboard(request):
#     if not request.user.is_super_admin:
#         messages.error(request, 'Access denied.')
#         return redirect('login')
    
#     # Import models
#     from schools.models import School, ReportingYear
#     from recycler_management.models import Recycler
#     from accounts.models import User
#     from waste_management.models import WasteEntry
    
#     # ========== OLD DATA (Jo pehle tha) ==========
#     total_schools = School.objects.count()
#     active_schools = School.objects.filter(status='active').count()
#     total_recyclers = Recycler.objects.count()
#     active_recyclers = Recycler.objects.filter(status='active').count()
#     total_users = User.objects.count()
#     recent_schools = School.objects.order_by('-created_at')[:5]
    
#     # ========== NEW ESG DATA (WasteEntry se) ==========
#     waste_entries = WasteEntry.objects.all()
    
#     # 1. Total Waste Generated
#     total_generated = waste_entries.aggregate(
#         total=Coalesce(Sum('weight_kg'), Decimal('0'))
#     )['total']
    
#     # 2. Total Diverted (Recycled + Composted + Co-processed)
#     diverted_entries = waste_entries.filter(
#         treatment_method__in=['recycled', 'composted', 'co_processed']
#     )
#     total_diverted = diverted_entries.aggregate(
#         total=Coalesce(Sum('weight_kg'), Decimal('0'))
#     )['total']
    
#     # 3. Diversion Rate
#     diversion_rate = 0
#     if total_generated > 0:
#         diversion_rate = (total_diverted / total_generated) * 100
    
#     # 4. CO₂e Avoided (sirf negative values matlab avoided)
#     co2e_avoided_kg = waste_entries.aggregate(
#         total=Coalesce(Sum('co2e_kg'), Decimal('0'))
#     )['total']
#     co2e_avoided_tonnes = co2e_avoided_kg / Decimal('1000')
    
#     # 5. Water Saved
#     water_saved = waste_entries.aggregate(
#         total=Coalesce(Sum('water_saved_litres'), Decimal('0'))
#     )['total']
    
#     # 6. Trees Saved
#     trees_saved = waste_entries.aggregate(
#         total=Coalesce(Sum('trees_equivalent'), Decimal('0'))
#     )['total']
    
#     # 7. Category-wise breakdown (Pie Chart ke liye)
#     from django.db.models import Sum as SumModel
#     category_breakdown = list(
#         waste_entries.values('waste_category')
#         .annotate(total_weight=Coalesce(SumModel('weight_kg'), Decimal('0')))
#         .order_by('-total_weight')[:6]
#     )
    
#     # 8. Monthly data (Bar Chart ke liye — last 6 months)
#     from django.db.models.functions import TruncMonth
#     from datetime import date, timedelta
#     from django.utils import timezone
    
#     six_months_ago = date.today() - timedelta(days=180)
#     monthly_data = list(
#         waste_entries.filter(entry_date__gte=six_months_ago)
#         .annotate(month=TruncMonth('entry_date'))
#         .values('month')
#         .annotate(
#             generated=Coalesce(SumModel('weight_kg'), Decimal('0')),
#             diverted=Coalesce(SumModel('weight_kg', filter=Q(treatment_method__in=['recycled', 'composted', 'co_processed'])), Decimal('0'))
#         )
#         .order_by('month')
#     )
    
#     # 9. Recent Waste Entries (Table ke liye)
#     recent_entries = waste_entries.select_related('reporting_year__school').order_by('-entry_date')[:10]
    
#     # 10. Reporting Years Stats
#     total_reporting_years = ReportingYear.objects.count()
#     active_reporting_years = ReportingYear.objects.filter(status='active').count()
    
#     # 11. Has entries or not
#     has_entries = waste_entries.exists()
    
#     context = {
#         # Old data
#         'total_schools': total_schools,
#         'active_schools': active_schools,
#         'total_recyclers': total_recyclers,
#         'active_recyclers': active_recyclers,
#         'total_users': total_users,
#         'recent_schools': recent_schools,
        
#         # New ESG KPIs
#         'total_generated': round(total_generated, 1),
#         'total_diverted': round(total_diverted, 1),
#         'diversion_rate': round(diversion_rate, 1),
#         'co2e_avoided': round(co2e_avoided_tonnes, 3),
#         'water_saved': round(water_saved, 0),
#         'trees_saved': round(trees_saved, 1),
        
#         # Charts data
#         'category_breakdown': category_breakdown,
#         'monthly_data': monthly_data,
        
#         # Tables
#         'recent_entries': recent_entries,
        
#         # Stats
#         'total_reporting_years': total_reporting_years,
#         'active_reporting_years': active_reporting_years,
#         'has_entries': has_entries,
#     }
    
#     return render(request, 'superadmin/dashboard.html', context)

@login_required(login_url='login')
def superadmin_dashboard(request):
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    # Import models
    from schools.models import School, ReportingYear
    from recycler_management.models import Recycler
    from accounts.models import User
    from waste_management.models import WasteEntry
    
    # ========== NEW CODE: GET SELECTED SCHOOL FROM FILTER ==========
    # User ne filter me school select kiya to uski ID aayegi
    selected_school_id = request.GET.get('school_id')
    selected_school = None
    
    # Agar user ne koi specific school select kiya hai ( 'all' nahi hai )
    if selected_school_id and selected_school_id != 'all':
        # Selected school ko database se fetch karo
        selected_school = School.objects.filter(id=selected_school_id).first()
        # Sirf us school ki waste entries filter karo
        waste_entries = WasteEntry.objects.filter(
            reporting_year__school=selected_school
        ) if selected_school else WasteEntry.objects.all()
    else:
        # Agar 'All Schools' select kiya hai to saari entries dikhao
        waste_entries = WasteEntry.objects.all()
        selected_school_id = 'all'
    # ===============================================================
    
    # ========== NEW CODE: All schools for filter dropdown ==========
    all_schools = School.objects.filter(status='active')
    # ===============================================================
    
    # ========== OLD DATA (Jo pehle tha - NO CHANGE) ==========
    total_schools = School.objects.count()
    active_schools = School.objects.filter(status='active').count()
    total_recyclers = Recycler.objects.count()
    active_recyclers = Recycler.objects.filter(status='active').count()
    total_users = User.objects.count()
    recent_schools = School.objects.order_by('-created_at')[:5]
    
    # ========== ESG DATA (Now uses filtered waste_entries) ==========
    # NOTE: waste_entries ab filter ke hisaab se filtered hai
    # Pehle: waste_entries = WasteEntry.objects.all()
    # Ab: waste_entries = filtered by selected school
    
    # 1. Total Waste Generated
    total_generated = waste_entries.aggregate(
        total=Coalesce(Sum('weight_kg'), Decimal('0'))
    )['total']
    
    # 2. Total Diverted (Recycled + Composted + Co-processed)
    diverted_entries = waste_entries.filter(
        treatment_method__in=['recycled', 'composted', 'co_processed']
    )
    total_diverted = diverted_entries.aggregate(
        total=Coalesce(Sum('weight_kg'), Decimal('0'))
    )['total']
    
    # 3. Diversion Rate
    diversion_rate = 0
    if total_generated > 0:
        diversion_rate = (total_diverted / total_generated) * 100
    
    # 4. CO₂e Avoided
    co2e_avoided_kg = waste_entries.aggregate(
        total=Coalesce(Sum('co2e_kg'), Decimal('0'))
    )['total']
    co2e_avoided_tonnes = co2e_avoided_kg / Decimal('1000')
    
    # 5. Water Saved
    water_saved = waste_entries.aggregate(
        total=Coalesce(Sum('water_saved_litres'), Decimal('0'))
    )['total']
    
    # 6. Trees Saved
    trees_saved = waste_entries.aggregate(
        total=Coalesce(Sum('trees_equivalent'), Decimal('0'))
    )['total']
    
    # 7. Category-wise breakdown (Pie Chart ke liye)
    from django.db.models import Sum as SumModel
    category_breakdown = list(
        waste_entries.values('waste_category')
        .annotate(total_weight=Coalesce(SumModel('weight_kg'), Decimal('0')))
        .order_by('-total_weight')[:6]
    )
    
    # 8. Monthly data (Bar Chart ke liye — last 6 months)
    from django.db.models.functions import TruncMonth
    from datetime import date, timedelta
    
    six_months_ago = date.today() - timedelta(days=180)
    monthly_data = list(
        waste_entries.filter(entry_date__gte=six_months_ago)
        .annotate(month=TruncMonth('entry_date'))
        .values('month')
        .annotate(
            generated=Coalesce(SumModel('weight_kg'), Decimal('0')),
            diverted=Coalesce(SumModel('weight_kg', filter=Q(treatment_method__in=['recycled', 'composted', 'co_processed'])), Decimal('0'))
        )
        .order_by('month')
    )
    
    # 9. Recent Waste Entries (Table ke liye)
    recent_entries = waste_entries.select_related('reporting_year__school').order_by('-entry_date')[:10]
    
    # 10. Reporting Years Stats
    total_reporting_years = ReportingYear.objects.count()
    active_reporting_years = ReportingYear.objects.filter(status='active').count()
    
    # 11. Has entries or not
    has_entries = waste_entries.exists()
    
    context = {
        # ========== NEW CODE: Filter data for template ==========
        'all_schools': all_schools,
        'selected_school_id': selected_school_id,
        'selected_school': selected_school,
        # =======================================================
        
        # Old data (NO CHANGE)
        'total_schools': total_schools,
        'active_schools': active_schools,
        'total_recyclers': total_recyclers,
        'active_recyclers': active_recyclers,
        'total_users': total_users,
        'recent_schools': recent_schools,
        
        # New ESG KPIs (NO CHANGE - but uses filtered data)
        'total_generated': round(total_generated, 1),
        'total_diverted': round(total_diverted, 1),
        'diversion_rate': round(diversion_rate, 1),
        'co2e_avoided': round(co2e_avoided_tonnes, 3),
        'water_saved': round(water_saved, 0),
        'trees_saved': round(trees_saved, 1),
        
        # Charts data (NO CHANGE - but uses filtered data)
        'category_breakdown': category_breakdown,
        'monthly_data': monthly_data,
        
        # Tables (NO CHANGE - but uses filtered data)
        'recent_entries': recent_entries,
        
        # Stats (NO CHANGE)
        'total_reporting_years': total_reporting_years,
        'active_reporting_years': active_reporting_years,
        'has_entries': has_entries,
    }
    
    return render(request, 'superadmin/dashboard.html', context)




@login_required
def change_password(request):
    """Force password change on first login"""
    
    # Check if user needs to change password
    if not request.user.must_change_password:
        messages.info(request, 'No password change required.')
        return redirect('superadmin_dashboard' if request.user.role == 'superadmin' else 'principal_dashboard')
    
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')
        
        # Check old password
        if not request.user.check_password(old_password):
            messages.error(request, 'Current password is incorrect.')
            return render(request, 'accounts/change_password.html')
        
        # Check new passwords match
        if new_password1 != new_password2:
            messages.error(request, 'New passwords do not match.')
            return render(request, 'accounts/change_password.html')
        
        # Check password length
        if len(new_password1) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
            return render(request, 'accounts/change_password.html')
        
        # Set new password
        request.user.set_password(new_password1)
        request.user.must_change_password = False
        request.user.save()
        
        # Keep user logged in
        update_session_auth_hash(request, request.user)
        
        messages.success(request, 'Password changed successfully!')
        return redirect('superadmin_dashboard' if request.user.role == 'superadmin' else 'principal_dashboard')
    
    return render(request, 'accounts/change_password.html')

# @login_required
# def principal_dashboard(request):
#     """Principal Dashboard"""
#     print(f"DEBUG: User = {request.user.username}")
#     print(f"DEBUG: User role = {request.user.role}")
    
#     if request.user.role != 'principal':
#         messages.error(request, 'Access denied.')
#         return redirect('login')
    
    
#     # Get principal profile
#     principal = Principal.objects.filter(user=request.user).first()
    
#     if not principal:
#         messages.error(request, 'Principal profile not found.')
#         return redirect('login')
    
#     school = principal.school
    
#     # Get active reporting year
#     active_year = ReportingYear.objects.filter(school=school, status='active').first()
    
#     # Get waste entries for this school (active year)
#     if active_year:
#         waste_entries = WasteEntry.objects.filter(reporting_year=active_year)
#     else:
#         waste_entries = WasteEntry.objects.none()
    
#     # Calculate KPIs
#     total_generated = waste_entries.aggregate(total=Sum('weight_kg'))['total'] or Decimal('0')
    
#     # Diverted (Recycled, Composted, Co-processed)
#     diverted_entries = waste_entries.filter(treatment_method__in=['recycled', 'composted', 'co_processed'])
#     total_diverted = diverted_entries.aggregate(total=Sum('weight_kg'))['total'] or Decimal('0')
    
#     # Diversion Rate
#     diversion_rate = 0
#     if total_generated > 0:
#         diversion_rate = (total_diverted / total_generated) * 100
    
#     # CO₂e Avoided
#     co2e_avoided_kg = waste_entries.filter(co2e_kg__lt=0).aggregate(total=Sum('co2e_kg'))['total'] or Decimal('0')
#     co2e_avoided_tonnes = abs(co2e_avoided_kg) / Decimal('1000')
    
#     # Water Saved
#     water_saved = waste_entries.aggregate(total=Sum('water_saved_litres'))['total'] or Decimal('0')
    
#     # Trees Saved
#     trees_saved = waste_entries.aggregate(total=Sum('trees_equivalent'))['total'] or Decimal('0')
    
#     # Recent entries
#     recent_entries = waste_entries.order_by('-entry_date')[:10]
    
#     context = {
#         'school': school,
#         'active_year': active_year,
#         'total_generated': round(total_generated, 1),
#         'total_diverted': round(total_diverted, 1),
#         'diversion_rate': round(diversion_rate, 1),
#         'co2e_avoided': round(co2e_avoided_tonnes, 2),
#         'water_saved': round(water_saved, 0),
#         'trees_saved': round(trees_saved, 1),
#         'recent_entries': recent_entries,
#         'has_entries': waste_entries.exists(),
#     }
    
#     # ========== YAHAN RENDER KARO, REDIRECT NAHI ==========
#     return render(request, 'principal/dashboard.html', context)

@login_required
def principal_dashboard(request):
    """Principal Dashboard with Charts"""
    print(f"DEBUG: User = {request.user.username}")
    print(f"DEBUG: User role = {request.user.role}")
    
    if request.user.role != 'principal':
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    # Get principal profile
    principal = Principal.objects.filter(user=request.user).first()
    
    if not principal:
        messages.error(request, 'Principal profile not found.')
        return redirect('login')
    
    school = principal.school
    
    # Get active reporting year
    active_year = ReportingYear.objects.filter(school=school, status='active').first()
    
    # Get waste entries for this school (active year)
    if active_year:
        waste_entries = WasteEntry.objects.filter(reporting_year=active_year)
    else:
        waste_entries = WasteEntry.objects.none()
    
    # ========== KPI CALCULATIONS ==========
    total_generated = waste_entries.aggregate(total=Sum('weight_kg'))['total'] or Decimal('0')
    
    # Diverted (Recycled, Composted, Co-processed)
    diverted_entries = waste_entries.filter(treatment_method__in=['recycled', 'composted', 'co_processed'])
    total_diverted = diverted_entries.aggregate(total=Sum('weight_kg'))['total'] or Decimal('0')
    
    # Diversion Rate
    diversion_rate = 0
    if total_generated > 0:
        diversion_rate = (total_diverted / total_generated) * 100
    
    # CO₂e Avoided
    co2e_avoided_kg = waste_entries.filter(co2e_kg__lt=0).aggregate(total=Sum('co2e_kg'))['total'] or Decimal('0')
    co2e_avoided_tonnes = abs(co2e_avoided_kg) / Decimal('1000')
    
    # Water Saved
    water_saved = waste_entries.aggregate(total=Sum('water_saved_litres'))['total'] or Decimal('0')
    
    # Trees Saved
    trees_saved = waste_entries.aggregate(total=Sum('trees_equivalent'))['total'] or Decimal('0')
    
    # ========== PIE CHART DATA (Category-wise breakdown) ==========
    from django.db.models import Sum as SumModel
    category_breakdown = list(
        waste_entries.values('waste_category')
        .annotate(total_weight=Coalesce(SumModel('weight_kg'), Decimal('0')))
        .order_by('-total_weight')[:6]
    )
    
    # ========== BAR CHART DATA (Monthly trend - last 6 months) ==========
    from django.db.models.functions import TruncMonth
    from datetime import date, timedelta
    
    six_months_ago = date.today() - timedelta(days=180)
    monthly_data = list(
        waste_entries.filter(entry_date__gte=six_months_ago)
        .annotate(month=TruncMonth('entry_date'))
        .values('month')
        .annotate(
            generated=Coalesce(SumModel('weight_kg'), Decimal('0')),
            diverted=Coalesce(SumModel('weight_kg', filter=Q(treatment_method__in=['recycled', 'composted', 'co_processed'])), Decimal('0'))
        )
        .order_by('month')
    )
    
    # Recent entries
    recent_entries = waste_entries.order_by('-entry_date')[:10]
    
    context = {
        'school': school,
        'active_year': active_year,
        'total_generated': round(total_generated, 1),
        'total_diverted': round(total_diverted, 1),
        'diversion_rate': round(diversion_rate, 1),
        'co2e_avoided': round(co2e_avoided_tonnes, 2),
        'water_saved': round(water_saved, 0),
        'trees_saved': round(trees_saved, 1),
        'recent_entries': recent_entries,
        'has_entries': waste_entries.exists(),
        # Chart data
        'category_breakdown': category_breakdown,
        'monthly_data': monthly_data,
    }
    
    return render(request, 'principal/dashboard.html', context)




def change_password_with_token(request, token):
    """Handle set password link from email - direct password set"""
    
    from accounts.models import User
    
    # Find user with this token
    user = User.objects.filter(set_password_token=token).first()
    
    if not user:
        messages.error(request, 'Invalid or expired link.')
        return redirect('login')
    
    # Check if token is still valid
    if not is_token_valid(user):
        messages.error(request, 'Link has expired. Please contact admin.')
        return redirect('login')
    
    # Handle password set form
    if request.method == 'POST':
        new_password1 = request.POST.get('new_password1')
        new_password2 = request.POST.get('new_password2')
        
        if new_password1 != new_password2:
            messages.error(request, 'Passwords do not match.')
        elif len(new_password1) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
        else:
            # Set password and activate account
            user.set_password(new_password1)
            user.is_active = True
            user.must_change_password = False
            user.set_password_token = None
            user.token_created_at = None
            user.save()
            
            # Log the user in
            login(request, user)
            
            messages.success(request, 'Password set successfully!')
            
            # Redirect based on role
            if user.role == 'principal':
                return redirect('principal_dashboard')
            else:
                return redirect('superadmin_dashboard')
    
    # Show password set form (your existing change_password.html)
    return render(request, 'accounts/change_password.html', {'email': user.email})


@login_required(login_url='login')
def teacher_dashboard(request):
    if request.user.role != 'teacher':
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    from waste_management.models import WasteEntry
    from django.db.models import Sum
    
    # Get teacher's school name
    school_name = request.user.school_name or "Your School"
    
    # Get teacher's waste entries
    entries = WasteEntry.objects.filter(teacher=request.user)
    
    # Calculate totals (they are Decimal type)
    total_entries = entries.count()
    total_weight = entries.aggregate(total=Sum('weight_kg'))['total'] or Decimal('0')
    total_co2 = entries.aggregate(total=Sum('co2e_kg'))['total'] or Decimal('0')
    total_water = entries.aggregate(total=Sum('water_saved_litres'))['total'] or Decimal('0')
    total_trees = entries.aggregate(total=Sum('trees_equivalent'))['total'] or Decimal('0')
    
    # ========== FIX: Convert Decimal to float for calculations ==========
    total_co2_float = float(total_co2)
    total_water_float = float(total_water)
    total_trees_float = float(total_trees)
    total_weight_float = float(total_weight)
    # ====================================================================
    
    # Calculate equivalents
    car_equivalent = round(total_co2_float / 4.6) if total_co2_float > 0 else 0
    shower_equivalent = round(total_water_float / 50) if total_water_float > 0 else 0
    
    # Recent entries (last 5)
    recent_entries = entries.order_by('-entry_date')[:5]
    
    context = {
        'school_name': school_name,
        'total_entries': total_entries,
        'total_weight': total_weight_float,
        'total_co2': total_co2_float,
        'total_water': total_water_float,
        'total_trees': total_trees_float,
        'recent_entries': recent_entries,
        'car_equivalent': car_equivalent,
        'shower_equivalent': shower_equivalent,
    }
    return render(request, 'teacher/dashboard.html', context)



@login_required(login_url='login')
def export_dashboard_pdf(request):
    """Export Super Admin Dashboard to PDF"""
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    # ========== SAME DATA AS DASHBOARD ==========

    
    # Get filter if applied
    selected_school_id = request.GET.get('school_id')
    selected_school = None
    
    if selected_school_id and selected_school_id != 'all':
        selected_school = School.objects.filter(id=selected_school_id).first()
        waste_entries = WasteEntry.objects.filter(reporting_year__school=selected_school) if selected_school else WasteEntry.objects.all()
    else:
        waste_entries = WasteEntry.objects.all()
    
    # Calculate KPIs
    total_generated = waste_entries.aggregate(total=Coalesce(Sum('weight_kg'), Decimal('0')))['total']
    
    diverted_entries = waste_entries.filter(treatment_method__in=['recycled', 'composted', 'co_processed'])
    total_diverted = diverted_entries.aggregate(total=Coalesce(Sum('weight_kg'), Decimal('0')))['total']
    
    diversion_rate = (total_diverted / total_generated * 100) if total_generated > 0 else 0
    
    co2e_avoided_kg = waste_entries.aggregate(total=Coalesce(Sum('co2e_kg'), Decimal('0')))['total']
    co2e_avoided_tonnes = co2e_avoided_kg / Decimal('1000')
    
    water_saved = waste_entries.aggregate(total=Coalesce(Sum('water_saved_litres'), Decimal('0')))['total']
    trees_saved = waste_entries.aggregate(total=Coalesce(Sum('trees_equivalent'), Decimal('0')))['total']
    
    # Category breakdown
    from django.db.models import Sum as SumModel
    category_breakdown = list(
        waste_entries.values('waste_category')
        .annotate(total_weight=Coalesce(SumModel('weight_kg'), Decimal('0')))
        .order_by('-total_weight')[:6]
    )
    
    # Recent entries
    recent_entries = waste_entries.select_related('reporting_year__school').order_by('-entry_date')[:10]
    
    # Stats
    total_schools = School.objects.count()
    total_recyclers = Recycler.objects.count()
    total_users = User.objects.count()
    
    context = {
        'total_generated': round(total_generated, 1),
        'total_diverted': round(total_diverted, 1),
        'diversion_rate': round(diversion_rate, 1),
        'co2e_avoided': round(co2e_avoided_tonnes, 3),
        'water_saved': round(water_saved, 0),
        'trees_saved': round(trees_saved, 1),
        'category_breakdown': category_breakdown,
        'recent_entries': recent_entries,
        'total_schools': total_schools,
        'total_recyclers': total_recyclers,
        'total_users': total_users,
        'selected_school': selected_school,
        'export_date': timezone.now(),
    }
    # ===========================================
    
    # Render HTML template
    template = get_template('superadmin/dashboard_pdf.html')
    html = template.render(context)
    
    # Create PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        # Set filename
        filename = f"ESG_Dashboard_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse('Error generating PDF', status=500)



@login_required(login_url='login')
def export_waste_excel(request):
    """Export Recent Waste Entries to Excel"""
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('login')

    # ===== SAME FILTER LOGIC AS DASHBOARD =====
    selected_school_id = request.GET.get('school_id')
    selected_school = None

    if selected_school_id and selected_school_id != 'all':
        selected_school = School.objects.filter(id=selected_school_id).first()
        waste_entries = WasteEntry.objects.filter(
            reporting_year__school=selected_school
        ) if selected_school else WasteEntry.objects.all()
    else:
        waste_entries = WasteEntry.objects.all()

    entries = waste_entries.select_related(
        'reporting_year__school'
    ).order_by('-entry_date')

    # ===== CREATE WORKBOOK =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waste Entries"

    # ----- Styles -----
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="1B5E20")  # dark green
    subhead_fill  = PatternFill("solid", fgColor="4CAF50")  # medium green
    subhead_font  = Font(bold=True, color="FFFFFF", size=10)
    center        = Alignment(horizontal="center", vertical="center")
    left          = Alignment(horizontal="left",   vertical="center")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ----- Title Row -----
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    school_label = selected_school.school_name if selected_school else "All Schools"
    title_cell.value = f"ESG Dashboard — Waste Entries Export  |  {school_label}"
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = PatternFill("solid", fgColor="1B5E20")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ----- Sub-title (generated date) -----
    ws.merge_cells("A2:F2")
    sub_cell = ws["A2"]
    from django.utils import timezone
    sub_cell.value = f"Generated: {timezone.now().strftime('%d %b %Y, %H:%M')}"
    sub_cell.font  = Font(italic=True, color="FFFFFF", size=9)
    sub_cell.fill  = PatternFill("solid", fgColor="2E7D32")
    sub_cell.alignment = center
    ws.row_dimensions[2].height = 18

    # ----- Column Headers -----
    headers = ["Date", "School", "Category", "Weight (kg)", "Treatment Method", "CO₂e (kg)"]
    for col_num, heading in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num, value=heading)
        cell.font      = subhead_font
        cell.fill      = subhead_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[3].height = 20

    # ----- Data Rows -----
    light_green = PatternFill("solid", fgColor="F1F8E9")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    for row_num, entry in enumerate(entries, start=4):
        row_fill = light_green if row_num % 2 == 0 else white_fill

        data = [
            entry.entry_date.strftime('%d %b %Y'),
            entry.reporting_year.school.school_name,
            entry.waste_category,
            float(entry.weight_kg),
            entry.get_treatment_method_display(),
            float(entry.co2e_kg),
        ]

        for col_num, value in enumerate(data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = center if col_num != 2 else left
        ws.row_dimensions[row_num].height = 16

    # ----- Column Widths -----
    col_widths = [14, 30, 22, 14, 22, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ----- Summary Row at bottom -----
    last_row = ws.max_row + 2
    ws.merge_cells(f"A{last_row}:C{last_row}")
    ws.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=1).alignment = center

    total_weight = sum(float(e.weight_kg) for e in entries)
    total_co2e   = sum(float(e.co2e_kg)   for e in entries)

    total_w_cell = ws.cell(row=last_row, column=4, value=round(total_weight, 2))
    total_w_cell.font = Font(bold=True)
    total_w_cell.fill = PatternFill("solid", fgColor="C8E6C9")
    total_w_cell.alignment = center

    total_c_cell = ws.cell(row=last_row, column=6, value=round(total_co2e, 3))
    total_c_cell.font = Font(bold=True)
    total_c_cell.fill = PatternFill("solid", fgColor="C8E6C9")
    total_c_cell.alignment = center

    # ===== SEND RESPONSE =====
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from django.utils import timezone
    filename = f"WasteEntries_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='login')
def export_category_excel(request):
    """Export Waste by Category to Excel"""
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('login')

    # ===== SAME FILTER LOGIC =====
    selected_school_id = request.GET.get('school_id')
    selected_school = None

    if selected_school_id and selected_school_id != 'all':
        selected_school = School.objects.filter(id=selected_school_id).first()
        waste_entries = WasteEntry.objects.filter(
            reporting_year__school=selected_school
        ) if selected_school else WasteEntry.objects.all()
    else:
        waste_entries = WasteEntry.objects.all()

    # Category breakdown - all categories (no limit)
    category_data = list(
        waste_entries
        .values('waste_category')
        .annotate(
            total_weight=Coalesce(Sum('weight_kg'), Decimal('0')),
            total_co2e=Coalesce(Sum('co2e_kg'), Decimal('0')),
            entry_count=Count('id'),
        )
        .order_by('-total_weight')
    )

    total_weight_all = sum(float(c['total_weight']) for c in category_data)

    # ===== CREATE WORKBOOK =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waste by Category"

    # ----- Styles -----
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1B5E20")
    subhead_fill = PatternFill("solid", fgColor="4CAF50")
    subhead_font = Font(bold=True, color="FFFFFF", size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ----- Title Row -----
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    school_label = selected_school.school_name if selected_school else "All Schools"
    title_cell.value = f"ESG Dashboard — Waste by Category  |  {school_label}"
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = PatternFill("solid", fgColor="1B5E20")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ----- Sub-title -----
    ws.merge_cells("A2:F2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {timezone.now().strftime('%d %b %Y, %H:%M')}  |  Total Categories: {len(category_data)}"
    sub_cell.font  = Font(italic=True, color="FFFFFF", size=9)
    sub_cell.fill  = PatternFill("solid", fgColor="2E7D32")
    sub_cell.alignment = center
    ws.row_dimensions[2].height = 18

    # ----- Column Headers -----
    headers = ["#", "Waste Category", "Total Weight (kg)", "% Share", "Total CO₂e (kg)", "No. of Entries"]
    for col_num, heading in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num, value=heading)
        cell.font      = subhead_font
        cell.fill      = subhead_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[3].height = 20

    # ----- Data Rows -----
    light_green = PatternFill("solid", fgColor="F1F8E9")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    for row_num, cat in enumerate(category_data, start=4):
        row_fill   = light_green if row_num % 2 == 0 else white_fill
        weight     = float(cat['total_weight'])
        co2e       = float(cat['total_co2e'])
        pct_share  = round((weight / total_weight_all * 100), 1) if total_weight_all > 0 else 0
        count      = cat['entry_count']

        row_data = [
            row_num - 3,                    # serial number
            cat['waste_category'],          # category name
            round(weight, 2),               # weight
            f"{pct_share}%",                # % share
            round(co2e, 3),                 # co2e
            count,                          # entries
        ]

        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = left if col_num == 2 else center
        ws.row_dimensions[row_num].height = 16

    # ----- Total Summary Row -----
    last_row = ws.max_row + 1
    total_fill = PatternFill("solid", fgColor="C8E6C9")
    bold_font  = Font(bold=True, size=10)

    ws.merge_cells(f"A{last_row}:B{last_row}")
    total_label = ws.cell(row=last_row, column=1, value="GRAND TOTAL")
    total_label.font      = bold_font
    total_label.fill      = total_fill
    total_label.alignment = center
    total_label.border    = border

    total_w = ws.cell(row=last_row, column=3, value=round(total_weight_all, 2))
    total_w.font = bold_font; total_w.fill = total_fill
    total_w.alignment = center; total_w.border = border

    pct_cell = ws.cell(row=last_row, column=4, value="100%")
    pct_cell.font = bold_font; pct_cell.fill = total_fill
    pct_cell.alignment = center; pct_cell.border = border

    total_co2e_all = sum(float(c['total_co2e']) for c in category_data)
    total_c = ws.cell(row=last_row, column=5, value=round(total_co2e_all, 3))
    total_c.font = bold_font; total_c.fill = total_fill
    total_c.alignment = center; total_c.border = border

    total_entries = sum(c['entry_count'] for c in category_data)
    total_e = ws.cell(row=last_row, column=6, value=total_entries)
    total_e.font = bold_font; total_e.fill = total_fill
    total_e.alignment = center; total_e.border = border

    ws.row_dimensions[last_row].height = 18

    # ----- Column Widths -----
    col_widths = [6, 28, 18, 12, 18, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== SEND RESPONSE =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"WasteByCategory_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='login')
def export_monthly_trend_excel(request):
    """Export Monthly Trend (Last 6 Months) to Excel"""
    if not request.user.is_super_admin:
        messages.error(request, 'Access denied.')
        return redirect('login')

    # ===== SAME FILTER LOGIC AS DASHBOARD =====
    selected_school_id = request.GET.get('school_id')
    selected_school = None

    if selected_school_id and selected_school_id != 'all':
        selected_school = School.objects.filter(id=selected_school_id).first()
        waste_entries = WasteEntry.objects.filter(
            reporting_year__school=selected_school
        ) if selected_school else WasteEntry.objects.all()
    else:
        waste_entries = WasteEntry.objects.all()

    # ===== CALCULATE MONTHLY DATA (LAST 6 MONTHS) =====
    from datetime import date, timedelta
    from django.db.models.functions import TruncMonth
    from django.db.models import Sum, Q
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    
    six_months_ago = date.today() - timedelta(days=180)
    monthly_data = list(
        waste_entries.filter(entry_date__gte=six_months_ago)
        .annotate(month=TruncMonth('entry_date'))
        .values('month')
        .annotate(
            generated=Coalesce(Sum('weight_kg'), Decimal('0')),
            diverted=Coalesce(Sum('weight_kg', filter=Q(treatment_method__in=['recycled', 'composted', 'co_processed'])), Decimal('0')),
            co2e=Coalesce(Sum('co2e_kg'), Decimal('0')),
        )
        .order_by('month')
    )

    # ===== CREATE WORKBOOK =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Trend"

    # ----- Styles -----
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1B5E20")
    subhead_fill = PatternFill("solid", fgColor="4CAF50")
    subhead_font = Font(bold=True, color="FFFFFF", size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ----- Title Row -----
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    school_label = selected_school.school_name if selected_school else "All Schools"
    title_cell.value = f"ESG Dashboard — Monthly Trend (Last 6 Months)  |  {school_label}"
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = PatternFill("solid", fgColor="1B5E20")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ----- Sub-title -----
    ws.merge_cells("A2:E2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {timezone.now().strftime('%d %b %Y, %H:%M')}  |  Total Months: {len(monthly_data)}"
    sub_cell.font  = Font(italic=True, color="FFFFFF", size=9)
    sub_cell.fill  = PatternFill("solid", fgColor="2E7D32")
    sub_cell.alignment = center
    ws.row_dimensions[2].height = 18

    # ----- Column Headers -----
    headers = ["Month", "Generated (kg)", "Diverted (kg)", "Diversion Rate", "CO₂e Saved (kg)"]
    for col_num, heading in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num, value=heading)
        cell.font      = subhead_font
        cell.fill      = subhead_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[3].height = 20

    # ----- Data Rows -----
    light_green = PatternFill("solid", fgColor="F1F8E9")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    for row_num, item in enumerate(monthly_data, start=4):
        row_fill = light_green if row_num % 2 == 0 else white_fill
        
        month_name = item['month'].strftime('%b %Y') if item['month'] else 'Unknown'
        generated = float(item['generated'])
        diverted = float(item['diverted'])
        diversion_rate = round((diverted / generated * 100), 1) if generated > 0 else 0
        co2e = float(item['co2e'])

        row_data = [
            month_name,
            round(generated, 1),
            round(diverted, 1),
            f"{diversion_rate}%",
            round(co2e, 2),
        ]

        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = center if col_num != 1 else left
        ws.row_dimensions[row_num].height = 16

    # ----- Summary Row -----
    last_row = ws.max_row + 1
    total_fill = PatternFill("solid", fgColor="C8E6C9")
    bold_font  = Font(bold=True, size=10)

    total_generated = sum(float(m['generated']) for m in monthly_data)
    total_diverted = sum(float(m['diverted']) for m in monthly_data)
    total_diversion_rate = round((total_diverted / total_generated * 100), 1) if total_generated > 0 else 0
    total_co2e = sum(float(m['co2e']) for m in monthly_data)

    ws.merge_cells(f"A{last_row}:A{last_row}")
    total_label = ws.cell(row=last_row, column=1, value="TOTAL")
    total_label.font      = bold_font
    total_label.fill      = total_fill
    total_label.alignment = center
    total_label.border    = border

    total_gen_cell = ws.cell(row=last_row, column=2, value=round(total_generated, 1))
    total_gen_cell.font = bold_font
    total_gen_cell.fill = total_fill
    total_gen_cell.alignment = center
    total_gen_cell.border = border

    total_div_cell = ws.cell(row=last_row, column=3, value=round(total_diverted, 1))
    total_div_cell.font = bold_font
    total_div_cell.fill = total_fill
    total_div_cell.alignment = center
    total_div_cell.border = border

    total_rate_cell = ws.cell(row=last_row, column=4, value=f"{total_diversion_rate}%")
    total_rate_cell.font = bold_font
    total_rate_cell.fill = total_fill
    total_rate_cell.alignment = center
    total_rate_cell.border = border

    total_co2_cell = ws.cell(row=last_row, column=5, value=round(total_co2e, 2))
    total_co2_cell.font = bold_font
    total_co2_cell.fill = total_fill
    total_co2_cell.alignment = center
    total_co2_cell.border = border

    ws.row_dimensions[last_row].height = 18

    # ----- Column Widths -----
    col_widths = [16, 18, 18, 16, 18]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== SEND RESPONSE =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"MonthlyTrend_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='login')
def export_principal_dashboard_pdf(request):
    """Export Principal Dashboard to PDF"""
    if request.user.role != 'principal':
        messages.error(request, 'Access denied.')
        return redirect('login')
    
    # Get principal profile
    from schools.models import Principal, ReportingYear
    from waste_management.models import WasteEntry
    from django.db.models import Sum
    from decimal import Decimal
    
    principal = Principal.objects.filter(user=request.user).first()
    
    if not principal:
        messages.error(request, 'Principal profile not found.')
        return redirect('login')
    
    school = principal.school
    
    # Get active reporting year
    active_year = ReportingYear.objects.filter(school=school, status='active').first()
    
    # Get waste entries for this school (active year)
    if active_year:
        waste_entries = WasteEntry.objects.filter(reporting_year=active_year)
    else:
        waste_entries = WasteEntry.objects.none()
    
    # Calculate KPIs
    total_generated = waste_entries.aggregate(total=Sum('weight_kg'))['total'] or Decimal('0')
    
    # Diverted (Recycled, Composted, Co-processed)
    diverted_entries = waste_entries.filter(treatment_method__in=['recycled', 'composted', 'co_processed'])
    total_diverted = diverted_entries.aggregate(total=Sum('weight_kg'))['total'] or Decimal('0')
    
    # Diversion Rate
    diversion_rate = 0
    if total_generated > 0:
        diversion_rate = (total_diverted / total_generated) * 100
    
    # CO₂e Avoided
    co2e_avoided_kg = waste_entries.filter(co2e_kg__lt=0).aggregate(total=Sum('co2e_kg'))['total'] or Decimal('0')
    co2e_avoided_tonnes = abs(co2e_avoided_kg) / Decimal('1000')
    
    # Water Saved
    water_saved = waste_entries.aggregate(total=Sum('water_saved_litres'))['total'] or Decimal('0')
    
    # Trees Saved
    trees_saved = waste_entries.aggregate(total=Sum('trees_equivalent'))['total'] or Decimal('0')
    
    # Recent entries
    recent_entries = waste_entries.order_by('-entry_date')[:10]
    
    context = {
        'school': school,
        'active_year': active_year,
        'total_generated': round(total_generated, 1),
        'total_diverted': round(total_diverted, 1),
        'diversion_rate': round(diversion_rate, 1),
        'co2e_avoided': round(co2e_avoided_tonnes, 2),
        'water_saved': round(water_saved, 0),
        'trees_saved': round(trees_saved, 1),
        'recent_entries': recent_entries,
        'has_entries': waste_entries.exists(),
        'export_date': timezone.now(),
    }
    
    # Render HTML template
    template = get_template('principal/dashboard_pdf.html')
    html = template.render(context)
    
    # Create PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        filename = f"Principal_Dashboard_{school.school_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse('Error generating PDF', status=500)

@login_required(login_url='login')
def export_principal_waste_excel(request):
    """Export Principal Recent Waste Entries to Excel"""
    if request.user.role != 'principal':
        messages.error(request, 'Access denied.')
        return redirect('login')

    from schools.models import Principal, ReportingYear

    principal = Principal.objects.filter(user=request.user).first()
    if not principal:
        messages.error(request, 'Principal profile not found.')
        return redirect('login')

    school = principal.school
    active_year = ReportingYear.objects.filter(school=school, status='active').first()

    if active_year:
        entries = WasteEntry.objects.filter(
            reporting_year=active_year
        ).order_by('-entry_date')
    else:
        entries = WasteEntry.objects.none()

    # ===== CREATE WORKBOOK =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waste Entries"

    # ----- Styles -----
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_white  = Font(bold=True, color="FFFFFF")
    light_green = PatternFill("solid", fgColor="F1F8E9")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    # ----- Title Row -----
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Waste Entries — {school.school_name}"
    title_cell.font      = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill      = PatternFill("solid", fgColor="1B5E20")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ----- Sub-title Row -----
    ws.merge_cells("A2:E2")
    sub_cell = ws["A2"]
    year_label = active_year.year if active_year else "No Active Year"
    sub_cell.value = f"Academic Year: {year_label}   |   Generated: {timezone.now().strftime('%d %b %Y, %H:%M')}"
    sub_cell.font      = Font(italic=True, color="FFFFFF", size=9)
    sub_cell.fill      = PatternFill("solid", fgColor="2E7D32")
    sub_cell.alignment = center
    ws.row_dimensions[2].height = 18

    # ----- Column Headers -----
    headers = ["Date", "Category", "Weight (kg)", "Treatment Method", "CO₂e (kg)"]
    for col_num, heading in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_num, value=heading)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="4CAF50")
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[3].height = 20

    # ----- Data Rows -----
    for row_num, entry in enumerate(entries, start=4):
        row_fill = light_green if row_num % 2 == 0 else white_fill

        row_data = [
            entry.entry_date.strftime('%d %b %Y'),
            entry.waste_category,
            float(entry.weight_kg),
            entry.get_treatment_method_display(),
            float(entry.co2e_kg),
        ]

        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = left if col_num == 2 else center
        ws.row_dimensions[row_num].height = 16

    # ----- Total Summary Row -----
    last_row = ws.max_row + 1
    total_fill = PatternFill("solid", fgColor="C8E6C9")
    bold_font  = Font(bold=True, size=10)

    ws.merge_cells(f"A{last_row}:B{last_row}")
    label_cell = ws.cell(row=last_row, column=1, value="GRAND TOTAL")
    label_cell.font = bold_font; label_cell.fill = total_fill
    label_cell.alignment = center; label_cell.border = border

    total_weight = sum(float(e.weight_kg) for e in entries)
    total_co2e   = sum(float(e.co2e_kg)   for e in entries)

    tw_cell = ws.cell(row=last_row, column=3, value=round(total_weight, 2))
    tw_cell.font = bold_font; tw_cell.fill = total_fill
    tw_cell.alignment = center; tw_cell.border = border

    ws.cell(row=last_row, column=4).fill = total_fill
    ws.cell(row=last_row, column=4).border = border

    tc_cell = ws.cell(row=last_row, column=5, value=round(total_co2e, 3))
    tc_cell.font = bold_font; tc_cell.fill = total_fill
    tc_cell.alignment = center; tc_cell.border = border

    ws.row_dimensions[last_row].height = 18

    # ----- Column Widths -----
    col_widths = [14, 28, 14, 24, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ===== SEND RESPONSE =====
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"WasteEntries_{school.school_name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response